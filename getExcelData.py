from openpyxl import Workbook, load_workbook

workbook =load_workbook(filename='FirstSheet.xlsx')
print('----------sheet names----------')
print(workbook.sheetnames)

print('--------active sheet-------')

sheet =workbook.active
print(sheet)
print('--------sheet title---------')
print(sheet.title)

print('--------sheet ["A1"]---------')
print(sheet['A1'])

print('--------sheet ["A1"].value---------')
print(sheet['A1'].value)

print('--------sheet.cell(row=10,col=1)---------')
print(sheet.cell(row=10,column=1))

print('--------sheet.cell(row=10,col=1).value---------')
print(sheet.cell(row=10,column=1).value)

print('---------iter_rows--------')

for row in sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=3):
    print(row)
print('--------iter_rows value-------')
for row in sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=3,values_only=True):
    print(row)


    print('------------iter_cols------------')

for row in sheet.iter_cols(min_row=1,max_row=2,min_col=1,max_col=3):
    print(row)
print('--------iter_rows value-------')
for row in sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=3,values_only=True):
    print(row)
